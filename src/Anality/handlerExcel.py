from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


class hanlderExcel:
    status = ''
    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')

    ft = Font(color="FFFFFFFF")

    def __init__(self, file_path):
        try:
            self.wb = load_workbook(file_path)
        except:
            self.status = 'Arquivo não encontrado!'

    def fill(self, sheet, list_to_fill):
        for item in list_to_fill:
            self.wb[sheet][item].fill = self.redFill
            self.wb[sheet][item].font = self.ft

    def salve(self, file_path):
        self.wb.save(file_path)


class excelToAnalyze(hanlderExcel):
    def __init__(self, file_path):
        super().__init__(file_path)


class excelTemplate(hanlderExcel):
    def __init__(self, file_path):
        super().__init__(file_path)
        self.list_cod = {
            str(row.value): row.row for row in self.wb['PARAMETROS']['A']
        }

    def findCod(self, cod):
        return (
            f'{self.list_cod[cod]}' if cod in self.list_cod.keys() else False
        )

    def creat_alert(self, infos):
        a = [row.value for row in self.wb['VERIFICAR_ERROS']['A']]
        n = a.index(None) + 1
        self.wb['VERIFICAR_ERROS'][f'A{n}'] = infos['COD']
        self.wb['VERIFICAR_ERROS'][f'B{n}'] = infos['ARQUIVO']
        self.wb['VERIFICAR_ERROS'][f'C{n}'] = infos['RAIO']
        self.wb['VERIFICAR_ERROS'][f'D{n}'] = infos['VALOR_KM']
        self.wb['VERIFICAR_ERROS'][f'E{n}'] = infos['INSTALACAO']
        self.wb['VERIFICAR_ERROS'][f'F{n}'] = infos['DESINSTALACAO']
        self.wb['VERIFICAR_ERROS'][f'G{n}'] = infos['POR_EQUIPAMENTO']
        self.wb['VERIFICAR_ERROS'][f'H{n}'] = infos['TOTAL_ERROS']
