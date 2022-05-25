from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font


class hanlderExcel:
    def __init__(self, file_path):
        self.wb = load_workbook(file_path)

    def fill(self, sheet, list_to_fill):
        my_fill = PatternFill(start_color='5399FF',
                              end_color='5399FF',
                              fill_type='solid')
        for item in list_to_fill:
            self.wb['sheet'][item].fill = my_fill
        
    def salve(self, file_path):
        self.wb.save(file_path)
    
class excelTemplate(hanlderExcel):
    def __init__(self, file_path):
        super().__init__(file_path)
        self.list_cod = {str(row.value): row.row for row in self.wb['PARAMETROS']['A']}
    
    def findCod(self, cod):
        return f'A{self.list_cod[cod]}' if cod in self.list_cod.keys() else False
